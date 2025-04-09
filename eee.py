import os
import cv2
import numpy as np
import mediapipe as mp
from deepface import DeepFace
from openpyxl import load_workbook, Workbook
import tkinter as tk
import threading

# تهيئة Mediapipe Face Detection
mp_face_detection = mp.solutions.face_detection
mp_drawing = mp.solutions.drawing_utils

# تحميل قاعدة بيانات الوجوه
def load_faces(folder_path):
    face_encodings = {}
    for person_name in os.listdir(folder_path):
        person_folder = os.path.join(folder_path, person_name)
        if os.path.isdir(person_folder):
            images = [os.path.join(person_folder, img) for img in os.listdir(person_folder) if
                      img.endswith(('.jpg', '.png'))]
            if images:
                face_encodings[person_name] = images[0]  # نأخذ صورة واحدة لكل شخص
    return face_encodings

# Paths to Excel files
excel_path_1 = "recognized_faces.xlsx"
excel_path_2 = "person_info.xlsx"
excel_path_3 = "rejected_names.xlsx"

# Function to create Excel files with headers
def create_excel_file(file_path, headers):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    wb.save(file_path)
    print(f"Created Excel file: {file_path}")

# Check if Excel files exist, create them if not
if not os.path.exists(excel_path_1):
    create_excel_file(excel_path_1, ["Name"])  # Header for recognized_faces.xlsx

if not os.path.exists(excel_path_2):
    create_excel_file(excel_path_2, ["Name", "Info1", "Info2", "Info3"])  # Example headers for person_info.xlsx

if not os.path.exists(excel_path_3):
    create_excel_file(excel_path_3, ["Name"])  # Header for rejected_names.xlsx

# Load Excel files
wb_1 = load_workbook(excel_path_1)
ws_1 = wb_1.active
wb_2 = load_workbook(excel_path_2)
ws_2 = wb_2.active

# Function to load rejected names from the third Excel file
def load_rejected_names(file_path):
    if not os.path.exists(file_path):
        create_excel_file(file_path, ["Name"])  # Create file if it doesn't exist
    wb = load_workbook(file_path)
    ws = wb.active
    return [row[0].value for row in ws.iter_rows(min_row=2, max_col=1) if row[0].value]

# Load rejected names dynamically
rejected_names = load_rejected_names(excel_path_3)

# Check if a name exists in the first Excel file
def is_name_exist(name):
    for row in ws_1.iter_rows(min_row=2, max_col=1):
        if row[0].value == name:
            return True
    return False

# Save a new name to the first Excel file
def save_name(name):
    if not is_name_exist(name):
        ws_1.append([name])
        wb_1.save(excel_path_1)
        print(f"Saved name: {name}")
        return True
    else:
        print(f"Name already exists: {name}")
        return False

# Get person information from the second Excel file
def get_person_info(name):
    for row in ws_2.iter_rows(min_row=2, max_col=ws_2.max_column):
        if row[0].value == name:
            return [cell.value for cell in row[1:]]  # Return all columns after the name
    return None

# تحديث نافذة المعلومات الرئيسية
def update_main_window(root, info, name):
    for widget in root.winfo_children():
        widget.destroy()

    # إنشاء الإطار الرئيسي
    main_frame = tk.Frame(root)
    main_frame.pack(fill=tk.BOTH, expand=True)

    # إطار المعلومات على اليسار
    info_frame = tk.Frame(main_frame)
    info_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)

    # تلوين الاسم حسب حالته (مرفوض أو غير ذلك)
    name_color = "red" if name in rejected_names else "green"
    name_label = tk.Label(info_frame, text=f"Name: {name}", font=("Arial", 14, "bold"), fg=name_color)
    name_label.pack(anchor="w", pady=5)

    if info:
        for idx, data in enumerate(info, start=1):
            tk.Label(info_frame, text=f"Info {idx}: {data}", font=("Arial", 12)).pack(anchor="w", pady=2)
    else:
        tk.Label(info_frame, text="No information available", font=("Arial", 12)).pack(anchor="w", pady=2)

    # إطار جانبي على اليمين
    sidebar_color = "red" if name in rejected_names else "green"  # أحمر للمرفوض، أخضر لغير ذلك
    sidebar_frame = tk.Frame(main_frame, width=100, bg=sidebar_color)
    sidebar_frame.pack(side=tk.RIGHT, fill=tk.Y)
    sidebar_frame.pack_propagate(False)  # منع تغيير حجم الإطار بناءً على المحتوى
    sidebar_text = "Warning" if name in rejected_names else "Safe"
    tk.Label(sidebar_frame, text=sidebar_text, bg=sidebar_color, fg="white", font=("Arial", 14)).pack(pady=20)

# التعرف على الوجوه في الوقت الفعلي
def recognize_faces(root):

    face_db = load_faces("database")
    cap = cv2.VideoCapture(0)

    with mp_face_detection.FaceDetection(min_detection_confidence=0.5) as face_detection:
        while cap.isOpened():
            ret, frame = cap.read()
            if not ret:
                break

            # تحويل الصورة إلى RGB
            rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

            # اكتشاف الوجوه
            results = face_detection.process(rgb_frame)

            if results.detections:
                for detection in results.detections:
                    bboxC = detection.location_data.relative_bounding_box
                    h, w, _ = frame.shape
                    x, y, w_box, h_box = int(bboxC.xmin * w), int(bboxC.ymin * h), int(bboxC.width * w), int(
                        bboxC.height * h)

                    face_crop = frame[y:y + h_box, x:x + w_box]
                    if face_crop.shape[0] > 0 and face_crop.shape[1] > 0:  # Check if face crop is valid
                        # مطابقة الوجه مع قاعدة البيانات باستخدام DeepFace
                        best_match = "unknown"
                        best_score = 0.3  # عتبة التعرف على الوجه (كلما قلّت القيمة، زادت الدقة)

                        for name, img_path in face_db.items():
                            try:
                                result = DeepFace.verify(face_crop, img_path, model_name='Facenet', enforce_detection=False)
                                if result["verified"] and result["distance"] < best_score:
                                    best_score = result["distance"]
                                    best_match = name
                            except Exception as e:
                                print(f"DeepFace verification error for {name}: {e}")

                        # رسم الصندوق واسم الشخص
                        box_color = (0, 0, 255) if best_match in rejected_names else (0, 255, 0)  # Red for rejected, green otherwise
                        cv2.rectangle(frame, (x, y), (x + w_box, y + h_box), box_color, 2)  # رسم صندوق حول الوجه
                        cv2.putText(frame, best_match, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, box_color, 2)  # كتابة الاسم

                        if best_match != "unknown":
                            is_new = save_name(best_match)
                            person_info = get_person_info(best_match)
                            update_main_window(root, person_info, best_match)
                        else:
                            update_main_window(root, None, "unknown")  # Show "No information available"

            cv2.imshow("Face Recognition", frame)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

    cap.release()
    cv2.destroyAllWindows()

# تشغيل البرنامج
if __name__ == "__main__":
    root = tk.Tk()
    root.title("Face Recognition System")
    root.geometry("800x600")  # ضبط حجم النافذة الرئيسية
    root.deiconify()  # التأكد من ظهور النافذة الرئيسية

    # تشغيل التعرف على الوجه في خيط مستقل
    face_thread = threading.Thread(target=recognize_faces, args=(root,), daemon=True)
    face_thread.start()

    root.mainloop()
